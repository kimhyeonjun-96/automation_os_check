from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font
from datetime import date
import re

from custom_modules.excel_modules.service.excelService import ExcelService

class HyudaiExcelServiceImpl(ExcelService):
    
    # 날짜
    tday = date.today()
    tday_s = tday.strftime("%Y-%m-%d")
    
    # 표지 스타일
    signSheet_border_style = Border(top=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'),
                    left=Side(border_style='thin', color='000000')
                    )
    signSheet_font_title_style = Font(name='맑은 고딕', size=25, color='000000', bold=True)
    signSheet_font_title2_style = Font(name='맑은 고딕', size=20, color='000000', bold=False)
    signSheet_font_nomal_style = Font(name='맑은 고딕', size=15, color='000000', bold=False)
    signSheet_alignment_title_style = Alignment(horizontal='center', vertical='center')
    signSheet_alignment_data2_style = alignment_data2_style = Alignment(horizontal='right', vertical='center')
    
    def __new__(cls, *args, **kwargs):
        if not hasattr(cls, "_instance"):
            cls._instance = super().__new__(cls)
        return cls._instance
    
    def __init__(self, savePath):
        cls = type(self)
        if not hasattr(cls, "_init"):
            cls._init = True
            self.savePath = savePath
            
    
    def createExcel(self):
        wb = Workbook()
        return wb
        
    def createSignSheet(self, wb):
        # ws = wb.active
        # ws.title = '표지'
        
        # 고객사 이름
        ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=9)
        ws['B2'] = '현대중공업'
        ws['B2'].font = self.signSheet_font_title_style
        ws['B2'].alignment = self.signSheet_alignment_title_style
        
        # 점검 일자
        ws.merge_cells(start_row=5, start_column=2, end_row=7, end_column=9)
        ws['B5'] = 'Linux S/W 점검'
        ws['B5'].font = self.signSheet_font_title2_style
        ws['B5'].alignment = self.signSheet_alignment_title_style
        
        ws.merge_cells(start_row=8, start_column=4, end_row=9, end_column=7)
        ws['D8'] = self.tday_s
        ws['D8'].font = self.signSheet_font_title2_style
        ws['D8'].alignment = self.signSheet_alignment_title_style
        
        # 서명 
        ws.merge_cells(start_row=14, start_column=2, end_row=15, end_column=5)
        ws.merge_cells(start_row=14, start_column=6, end_row=15, end_column=9)
        ws['B14'] = '점검자'
        ws['F14'] = '확인자'
        
        ws['B14'].font = self.signSheet_font_title2_style
        ws['F14'].font = self.signSheet_font_title2_style
        
        ws['B14'].alignment = self.signSheet_alignment_title_style
        ws['F14'].alignment = self.signSheet_alignment_title_style
        
        ws['B14'].border = self.signSheet_border_style
        ws['F14'].border = self.signSheet_border_style
        
        ws.merge_cells(start_row=16, start_column=2, end_row=17, end_column=5)
        ws.merge_cells(start_row=16, start_column=6, end_row=17, end_column=9) 
        ws['B16'] = '(인)'
        ws['F16'] = '(인)'
        
        ws['B16'].font = self.signSheet_font_nomal_style
        ws['F16'].font = self.signSheet_font_nomal_style
        
        ws['B16'].alignment = self.signSheet_alignment_data2_style
        ws['F16'].alignment = self.signSheet_alignment_data2_style
        
        ws['B16'].border = self.signSheet_border_style
        ws['F16'].border = self.signSheet_border_style
        
    def createCheckItemSheet(self, wb):
        ws = wb.active
        ws.title = "점검항목"
        # ws = wb.create_sheet('점검항목')
        
        # 첫 시트 데이터 지정
        # 표, 폰트, 색, 정렬
        border_style = Border(top=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000')
                            )
        
        font_title_style = Font(name='맑은 고딕', size=15, color='000000', bold=True)
        font_normal_style = Font(name='맑은 고딕', size=12, color='000000', bold=False)
        
        alignment_title_style = Alignment(horizontal='center', vertical='center')
        alignment_data_style = Alignment(horizontal='center', vertical='center',wrap_text=True)
        alignment_data2_style = Alignment(horizontal='left', vertical='center')
        alignment_data3_style = Alignment(horizontal='center', vertical='center',wrap_text=True)
        

        fill_style = PatternFill(start_color='CCCCFF', fill_type = 'solid')
        # 1행 머지 , 년 월 정기점검 데이터 입력 , 중앙정렬, 색 입히기
        ws.merge_cells('A1:O1')
        tday = date.today()
        tday_s = tday.strftime("%Y년 %m월")
        ws['A1'] = tday_s + ' 정기점검'
        ws['A1'].font = font_title_style
        ws['A1'].alignment = alignment_title_style
        
        
        
        # 2행 (abc 머지{고객사 데이터 입력, 중앙정렬, 색 입기},f{점검기간 데이터 입력, 중앙정렬, 색 입히기} de 머지, gh 머지, ijk 머지{담당자 데이터 입력, 중앙정렬, 색 입히기}, LMNO 머지)
        ws.merge_cells('A2:C2')
        ws.merge_cells('D2:E2')
        ws.merge_cells('G2:H2')
        ws.merge_cells('I2:K2')
        ws.merge_cells('L2:O2')
        ws['A2'] = '고객사'
        ws['A2'].font = font_normal_style
        ws['A2'].alignment = alignment_data_style
        ws['A2'].fill = fill_style
        
        ws['F2'] = '점검기간'
        ws['F2'].font = font_normal_style
        ws['F2'].alignment = alignment_data_style
        ws['F2'].fill = fill_style
        
        ws['D2'].font = font_normal_style
        ws['D2'].alignment = alignment_data_style
        
        ws['G2'].font = font_normal_style
        ws['G2'].alignment = alignment_data_style
        
        ws['I2'] = '담당자'
        ws['I2'].font = font_normal_style
        ws['I2'].alignment = alignment_data_style
        ws['I2'].fill = fill_style
        
        ws['L2'].font = font_normal_style
        ws['L2'].alignment = alignment_data_style
        
        # 3,4행 (abc 머지{OS 데이터 입력, 중앙정렬, 색 입히기}, defgh 머지, ijk 머지{락플레이스 담당자 데이터 입력, 중앙정렬, 색 입히기}, LMNO 머지)
        ws.merge_cells('A3:C3')
        ws.merge_cells('D3:H3')
        ws.merge_cells('I3:K3')
        ws.merge_cells('L3:O3')
        
        ws.merge_cells('A4:C4')
        ws.merge_cells('D4:H4')
        ws.merge_cells('I4:K4')
        ws.merge_cells('L4:O4')
        
        ws['A3'] = 'OS'
        ws['A3'].font = font_normal_style
        ws['A3'].alignment = alignment_data_style
        ws['A3'].fill = fill_style
        
        ws['D3'].font = font_normal_style
        ws['D3'].alignment = alignment_data_style
        
        ws['I3'] = '락플레이스 점검자'
        ws['I3'].font = font_normal_style
        ws['I3'].alignment = alignment_data_style
        ws['I3'].fill = fill_style
        
        ws['L3'].font = font_normal_style
        ws['L3'].alignment = alignment_data_style
        
        ws['A4'] = '총 수량'
        ws['A4'].font = font_normal_style
        ws['A4'].alignment = alignment_data_style
        ws['A4'].fill = fill_style
        
        ws['D4'].font = font_normal_style
        ws['D4'].alignment = alignment_data_style
        
        ws['I4'] = '락플레이스 영업대표'
        ws['I4'].font = font_normal_style
        ws['I4'].alignment = alignment_data_style
        ws['I4'].fill = fill_style
        
        
        ws['L4'].font = font_normal_style
        ws['L4'].alignment = alignment_data_style
        
        # 5행 전체 머지
        ws.merge_cells('A5:O5')
        
        # 6~34행 (ab 머지{6행 구분 데이터 입력 / 7~34행 os 데이터 입력, 중앙정렬, 색입히기}, cde머지{점검항목 데이터 입력, 중앙정렬, 색입히기}, fghijk 머지{점검 방법 데이터 입력, 중앙정렬, 색 입히기}, LMNO 머지{점검기준 데이터 입력, 중앙정렬, 색 입히기})
        for i in range(6,19):
            if i == 6:
                ws.merge_cells('A'+str(i)+':'+'B'+str(i))
            
            if i != 15 and i != 16 and i != 17 : 
                ws.merge_cells('C'+str(i)+':'+'E'+str(i))
                ws.merge_cells('L'+str(i)+':'+'O'+str(i))
            
            ws.merge_cells('F'+str(i)+':'+'K'+str(i))
            
        ws['A6'] = '구  분'
        ws['A6'].font = font_normal_style
        ws['A6'].alignment = alignment_data_style
        ws['A6'].fill = fill_style
        
        ws['C6'] = '점  검  항  목'
        ws['C6'].font = font_normal_style
        ws['C6'].alignment = alignment_data_style
        ws['C6'].fill = fill_style
        
        ws['F6'] = '점  검  방  법'
        ws['F6'].font = font_normal_style
        ws['F6'].alignment = alignment_data_style
        ws['F6'].fill = fill_style
        
        ws['L6'] = '점  검  기  준'
        ws['L6'].font = font_normal_style
        ws['L6'].alignment = alignment_data_style
        ws['L6'].fill = fill_style
        
        ws.merge_cells(start_row=7, start_column=1, end_row=18, end_column=2)
        ws['A7'] = 'OS'
        ws['A7'].font = font_normal_style
        ws['A7'].alignment = alignment_data_style
        ws['A7'].fill = fill_style
        
        ws.merge_cells(start_row=15, start_column=3, end_row=17, end_column=5)

        ws['C7'] = 'Hostname'
        ws['C7'].font = font_normal_style
        ws['C7'].alignment = alignment_data_style
            
        ws['C8'] = 'OS Release'
        ws['C8'].font = font_normal_style
        ws['C8'].alignment = alignment_data_style
        
        ws['C9'] = 'Kernel Release'
        ws['C9'].font = font_normal_style
        ws['C9'].alignment = alignment_data_style
        
        ws['C10'] = 'CPU'
        ws['C10'].font = font_normal_style
        ws['C10'].alignment = alignment_data_style
        
        ws['C11'] = 'Memory'
        ws['C11'].font = font_normal_style
        ws['C11'].alignment = alignment_data_style
        
        ws['C12'] = 'Disk'
        ws['C12'].font = font_normal_style
        ws['C12'].alignment = alignment_data_style
        
        ws['C13'] = 'Process'
        ws['C13'].font = font_normal_style
        ws['C13'].alignment = alignment_data_style
        
        ws['C14'] = 'System Log'
        ws['C14'].font = font_normal_style
        ws['C14'].alignment = alignment_data_style
        
        ws['C15'] = 'Network'
        ws['C15'].font = font_normal_style
        ws['C15'].alignment = alignment_data_style
        
        ws['C18'] = 'Uptime'
        ws['C18'].font = font_normal_style
        ws['C18'].alignment = alignment_data_style
    
        ws.merge_cells(start_row=15, start_column=12, end_row=17, end_column=15)

        ws['F7'] = '# hostname'
        ws['F7'].font = font_normal_style
        ws['F7'].alignment = alignment_data2_style
            
        ws['F8'] = '# cat /etc/redhat-release'
        ws['F8'].font = font_normal_style
        ws['F8'].alignment = alignment_data2_style
        
        ws['F9'] = '# uname -r'
        ws['F9'].font = font_normal_style
        ws['F9'].alignment = alignment_data2_style
        
        ws['F10'] = '# cat /proc/cpuinfo'
        ws['F10'].font = font_normal_style
        ws['F910'].alignment = alignment_data2_style
        
        ws['F11'] = '# cat /proc/meminfo'
        ws['F11'].font = font_normal_style
        ws['F11'].alignment = alignment_data2_style
        
        ws['F12'] = '# df -h'
        ws['F12'].font = font_normal_style
        ws['F12'].alignment = alignment_data2_style
        
        ws['F13'] = '# pstree # ps aux'
        ws['F13'].font = font_normal_style
        ws['F13'].alignment = alignment_data2_style
        
        ws['F14'] = '# dmesg # cat /var/log/messages'
        ws['F14'].font = font_normal_style
        ws['F14'].alignment = alignment_data2_style
        
        ws['F15'] = '# ifconfig'
        ws['F15'].font = font_normal_style
        ws['F15'].alignment = alignment_data2_style
        
        ws['F16'] = '# ethtool (nic) , # netstat -anuten'
        ws['F16'].font = font_normal_style
        ws['F16'].alignment = alignment_data2_style
        
        ws['F17'] = '# cat /proc/net/bonding/(bonding)'
        ws['F17'].font = font_normal_style
        ws['F17'].alignment = alignment_data2_style
        
        ws['F18'] = '# uptime'
        ws['F18'].font = font_normal_style
        ws['F18'].alignment = alignment_data2_style
        
        ws['L7'] = 'hostname 확인'
        ws['L7'].font = font_normal_style
        ws['L7'].alignment = alignment_data_style
            
        ws['L8'] = 'RHEL Release 확인'
        ws['L8'].font = font_normal_style
        ws['L8'].alignment = alignment_data_style
        
        ws['L9'] = 'Kernel Release 확인'
        ws['L9'].font = font_normal_style
        ws['L9'].alignment = alignment_data_style
        
        ws['L10'] = 'CPU 정보 확인'
        ws['L10'].font = font_normal_style
        ws['L10'].alignment = alignment_data_style
   
        ws['L11'] = '메모리 크기의 80% 이하'
        ws['L11'].font = font_normal_style
        ws['L11'].alignment = alignment_data_style
        
        ws['L12'] = '파티션 크기의 80% 이하'
        ws['L12'].font = font_normal_style
        ws['L12'].alignment = alignment_data_style
        
        ws['L13'] = 'Zombie 프로세스 여부'
        ws['L13'].font = font_normal_style
        ws['L13'].alignment = alignment_data_style
        
        ws['L14'] = '특이 로그 확인'
        ws['L14'].font = font_normal_style
        ws['L14'].alignment = alignment_data_style
        
        ws['L15'] = 'VM 자체 네트워크' + ' 이중화 구성으로 점검 제외'
        ws['L15'].font = font_normal_style
        ws['L15'].alignment = alignment_data3_style
        
        ws['L18'] = '180 days 이하, Core 당 1 미만'
        ws['L18'].font = font_normal_style
        ws['L18'].alignment = alignment_data_style
        
        # 22~24행 전체 머지 {점검자:   / 담당자:   데이터 입력, 아래중앙정렬}
        ws.merge_cells(start_row=19, start_column=1, end_row=21, end_column=15)
        
        ws['A19'] = '점 검 자:                                        담 당 자:                                        '
        ws['A19'].font = font_normal_style
        ws['A19'].alignment = alignment_data_style
        
        # 전체테두리
        # for cells in ws.columns:
        #     for cell in cells:
        #         print(cell)
        #         ws[cell.coordinate].border = border_style
        for num, row in enumerate(ws.rows):
            for cell in row:
                # print(num)
                if num < 21:
                    cell.border = border_style
         
    def createCheckSheet(self,wb):
        ws = wb.create_sheet('점검리스트' )
        
        ws['A1'] = 'hostname'
        ws['B1'] = 'os-version'
        ws['C1'] = 'kernel-version'
        ws['D1'] = 'uptime'
        ws['E1'] = 'cpu'
        ws['F1'] = 'mem'
        ws['G1'] = 'swap'
        ws['H1'] = 'disk'
        ws['I1'] = 'network'
        ws['J1'] = 'zombie'
        ws['K1'] = 'ntp'
        ws['L1'] = 'log'
        
        return ws
    
    def saveExcel(self, wb):
        fileName = self.savePath +"/RP_OS_CHECK_" + self.tday_s + ".xlsx"
        wb.save(fileName)
    
    def checkResource(self, ws):
        fill_style = PatternFill(start_color='FFFF00', fill_type = 'solid')
        
        for cells in ws:
            for cell in cells:
                if cell != cells[0] and cell.value: 
                    if 'B' in cell.coordinate:
                        if "Failed to connect to the host via ssh" in cell.value:
                            ws[cell.coordinate].fill = fill_style
                    if 'D' in cell.coordinate or 'E' in cell.coordinate or 'F' in cell.coordinate or 'G' in cell.coordinate:
                        if ' days' in cell.value:
                            day = int(cell.value.replace(" days", ""))
                            if day >= 80:
                                ws[cell.coordinate].fill = fill_style
                        elif '%' in cell.value:
                            uses = int(float(re.sub(r'[^0-9.]', '' ,cell.value)))
                            if uses >= 80:
                                ws[cell.coordinate].fill = fill_style
                    elif  'H' in cell.coordinate:
                        if '%' in cell.value:
                            data_list = cell.value.split('\n')
                            del data_list[-1]
                            for data in data_list:
                                uses = int(float(re.sub(r'[^0-9.]', '' ,data)))
                                if uses >= 80:
                                    ws[cell.coordinate].fill = fill_style
                    elif 'J' in cell.coordinate:
                        if 'zombie' not in cell.value and int(cell.value) > 0:
                            ws[cell.coordinate].fill = fill_style
        
        for cells in ws.columns:
            for cell in cells:
                if cell != cells[0] and cell.value: 
                    if 'I' in cell.coordinate or 'K' in cell.coordinate or 'L' in cell.coordinate:
                        if 'fail' in cell.value:
                            ws[cell.coordinate].fill = fill_style
                        elif 'pass' not in cell.value:
                            ws[cell.coordinate].fill = fill_style
        
        
    def createCustomSheet(self):
        print('createCustomSheet')
    
    def writeData(self, ws, host_list):
        for host in host_list:
            if host.getSshMsg() =='':
                ws.append([host.getHostname(), host.getVersion(), host.getKernel(), host.getUptime(), host.getUseCpu(), host.getUseRealMem(), host.getUseSwapMem(), host.getUseDisk(), host.getNetwork(), host.getZom(), host.getNtp(), host.getLog()])
            else:
                ws.append([host.getHostname(), host.getSshMsg()])
    
    def dataStyle(self, ws):
        font_title_style = Font(name='맑은 고딕', size=15, color='000000', bold=True)
        font_normal_style = Font(name='맑은 고딕', size=12, color='000000', bold=False)

        alignment_title_style = Alignment(horizontal='center', vertical='center')
        alignment_hostname_style = Alignment(horizontal='center', vertical='center',wrap_text=True)
        alignment_data_style = Alignment(horizontal='center', vertical='center')
        alignment_data2_style = Alignment(horizontal='left', vertical='center',wrap_text=True)

        border_style = Border(top=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000')
                            )
        fill_style = PatternFill(start_color='CCCCFF', fill_type = 'solid')
        
        # 데이터 셀 중앙 정렬 , 데이터 셀 폰트 설정
        for cells in ws.columns:
            for cell in cells:
                if cell.coordinate != cells[0] and cell.value:
                    if len(cell.value) >= 10:
                        ws[cell.coordinate].alignment = alignment_data2_style
                    else:
                        ws[cell.coordinate].alignment = alignment_data_style
                    ws[cell.coordinate].font = font_normal_style
        
        # check_sheet 타이틀 폰트, 셀 색, 가운데 정렬, 
        for cells in ws.columns:    
            ws[cells[0].coordinate].font = font_title_style
            ws[cells[0].coordinate].fill = fill_style
            ws[cells[0].coordinate].alignment = alignment_title_style

                
        # check_sheet 호스트네임 색 채우기, 가운데 정렬 
        for cells in ws.rows:
            ws[cells[0].coordinate].fill = fill_style
            ws[cells[0].coordinate].alignment = alignment_hostname_style
        
        # 전체테두리
        for cells in ws.columns:
            for cell in cells:
                ws[cell.coordinate].border = border_style
        
        for column_cells in ws.columns:
            if column_cells[0].column_letter == 'H' or column_cells[0].column_letter == 'I':
                ws.column_dimensions[column_cells[0].column_letter].width = 45
            else:
                length = max(len(str(cell.value))*1.5 for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length
                ws[column_cells[0].coordinate].alignment = alignment_title_style
        
        for cells in ws.rows:
            ws[cells[1].coordinate].alignment = alignment_hostname_style
            ws[cells[2].coordinate].alignment = alignment_hostname_style
            
    def excel_start(self, host_list):        
        # 엑셀 생성
        custom_wb = self.createExcel()
        
        # 시트 생성
        # self.createSignSheet(custom_wb)
        self.createCheckItemSheet(custom_wb)
        check_ws = self.createCheckSheet(custom_wb)
        
        
        self.writeData(check_ws, host_list)
        self.dataStyle(check_ws)
        self.checkResource(check_ws)
        
        # 엑셀 저장
        self.saveExcel(custom_wb)
        

        
    